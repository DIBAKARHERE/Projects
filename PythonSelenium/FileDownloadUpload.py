import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

file_path = "C:\\Users\\AZURE\\Downloads\\download.xlsx"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.maximize_window()
Dict={}

# driver.find_element(By.CSS_SELECTOR, "#downloadButton").click()
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)   # to select for upload file we have to pass the file path inside send_keys

wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text) # **** driver.find_element expects 2 arguments but toast_locator is a tuple. '*' unpacks the tuple by extracting 2 elements which is expected

# print(driver.find_element(By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)").text)

book = openpyxl.load_workbook(file_path)
sheet = book.active

# for i in range(1,sheet.max_row+1):
#     if sheet.cell(row=i, column=1).value == 2:   #we match the specific value to print the specific row
#         for j in range(1,sheet.max_column+1):
#             print(sheet.cell(row=i, column=j).value)   #to pull up the entire row data

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == 1:   #we match the specific value to print the specific row

            for j in range(2,sheet.max_column+1):
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)









