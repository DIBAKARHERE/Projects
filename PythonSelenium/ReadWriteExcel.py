import openpyxl
book = openpyxl.load_workbook("C:\\Users\\AZURE\\Downloads\\databook.xlsx")
sheet = book.active   #The value of the active sheet will be printed for the current sheet is saved
# sheet = book.worksheets[0]  #If want to print the value of the specific sheet we can do indexing
cell = sheet.cell(row=2, column=2)
print(cell.value)

sheet.cell(row=1, column=5).value = "Final Status"
print(sheet.cell(row=1, column=5).value)
book.save("C:\\Users\\AZURE\\Downloads\\databook.xlsx") #"/" -> URL separator, "\\" -> for defining path

print(sheet.max_row)
print(sheet.max_column)
# print(sheet['B5'].value)     #another way to print specific cell by indexing the cell number

for i in range(2, sheet.max_row+1):
    for j in range(2, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)